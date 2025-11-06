import os
import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.conf import settings
from app_clients_resumes.models import Resume, TutorProfile


class Command(BaseCommand):
    help = 'Импорт резюме из Excel файлов'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            help='Путь к Excel файлу для импорта',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Показать что будет импортировано без сохранения в БД',
        )

    def handle(self, *args, **options):
        file_path = options.get('file')
        dry_run = options.get('dry_run', False)

        # Если файл не указан, ищем все Excel файлы в static/xlsx-files/
        if not file_path:
            xlsx_dir = os.path.join(settings.BASE_DIR, 'static', 'xlsx-files')
            if not os.path.exists(xlsx_dir):
                raise CommandError(f'Папка {xlsx_dir} не существует')
            
            excel_files = [f for f in os.listdir(xlsx_dir) if f.endswith('.xlsx')]
            if not excel_files:
                raise CommandError(f'В папке {xlsx_dir} не найдено Excel файлов')
            
            self.stdout.write(f'Найдено {len(excel_files)} Excel файлов для обработки')
            
            for excel_file in excel_files:
                file_full_path = os.path.join(xlsx_dir, excel_file)
                self.process_excel_file(file_full_path, dry_run)
        else:
            if not os.path.exists(file_path):
                raise CommandError(f'Файл {file_path} не существует')
            self.process_excel_file(file_path, dry_run)

        if dry_run:
            self.stdout.write(
                self.style.WARNING('Это был пробный запуск. Данные не были сохранены в БД.')
            )
        else:
            self.stdout.write(
                self.style.SUCCESS('Импорт резюме завершен успешно!')
            )

    def process_excel_file(self, file_path, dry_run):
        """Обрабатывает один Excel файл"""
        self.stdout.write(f'\nОбработка файла: {os.path.basename(file_path)}')
        
        try:
            workbook = openpyxl.load_workbook(file_path)
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f'Ошибка при открытии файла {file_path}: {e}')
            )
            return

        # Обрабатываем все листы
        for sheet_name in workbook.sheetnames:
            self.stdout.write(f'  Обработка листа: {sheet_name}')
            worksheet = workbook[sheet_name]
            self.process_worksheet(worksheet, dry_run, sheet_name)

    def process_worksheet(self, worksheet, dry_run, sheet_name):
        """Обрабатывает один лист Excel файла"""
        if worksheet.max_row < 2:
            self.stdout.write(f'    Лист {sheet_name} пуст или содержит только заголовки')
            return

        # Получаем заголовки из первой строки
        headers = []
        for col in range(1, worksheet.max_column + 1):
            header = worksheet.cell(row=1, column=col).value
            headers.append(header)

        self.stdout.write(f'    Найдено столбцов: {len(headers)}')
        self.stdout.write(f'    Заголовки: {headers[:8]}...')  # Показываем первые 8

        # Определяем индексы столбцов с резюме
        resume_columns = self.identify_resume_columns(headers)
        if not resume_columns:
            self.stdout.write(f'    В листе {sheet_name} не найдено столбцов с резюме')
            return

        self.stdout.write(f'    Найдено столбцов с резюме: {len(resume_columns)}')

        # Обрабатываем каждую строку данных
        processed_count = 0
        saved_count = 0
        
        for row_num in range(2, worksheet.max_row + 1):
            # Получаем ID ребенка (первый столбец)
            student_id_cell = worksheet.cell(row=row_num, column=1)
            if student_id_cell.value is None:
                continue
            
            # Правильно обрабатываем числовые ID (убираем .0 если это целое число)
            if isinstance(student_id_cell.value, (int, float)):
                if float(student_id_cell.value).is_integer():
                    student_crm_id = str(int(student_id_cell.value))
                else:
                    student_crm_id = str(student_id_cell.value)
            else:
                student_crm_id = str(student_id_cell.value).strip()
            
            if not student_crm_id or student_crm_id == 'None':
                continue

            # Получаем ФИО ребенка (второй столбец) для логирования
            student_name_cell = worksheet.cell(row=row_num, column=2)
            student_name = str(student_name_cell.value).strip() if student_name_cell.value else 'Неизвестно'

            processed_count += 1

            # Обрабатываем каждый столбец с резюме
            for col_index, resume_header in resume_columns.items():
                resume_cell = worksheet.cell(row=row_num, column=col_index)
                resume_content = resume_cell.value
                
                # Пропускаем пустые резюме
                if not resume_content or str(resume_content).strip() in ['', 'None']:
                    continue

                resume_content = str(resume_content).strip()
                
                if dry_run:
                    self.stdout.write(
                        f'    [DRY RUN] Сохранил бы резюме: ID={student_crm_id}, '
                        f'Имя={student_name}, Длина={len(resume_content)} символов'
                    )
                else:
                    # Сохраняем резюме в БД
                    resume, created = Resume.objects.get_or_create(
                        student_crm_id=student_crm_id,
                        content=resume_content,
                        defaults={
                            'is_verified': False
                        }
                    )
                    
                    if not created:
                        # Обновляем существующее резюме
                        resume.content = resume_content
                        resume.save()
                        action = 'обновлено'
                    else:
                        action = 'создано'
                    
                    saved_count += 1
                    self.stdout.write(
                        f'    Резюме {action}: ID={student_crm_id}, '
                        f'Имя={student_name}'
                    )

        self.stdout.write(
            f'    Обработано строк: {processed_count}, '
            f'Сохранено резюме: {saved_count}'
        )

    def identify_resume_columns(self, headers):
        """Определяет какие столбцы содержат резюме"""
        resume_columns = {}
        
        for i, header in enumerate(headers, 1):
            if not header:
                continue
                
            header_lower = str(header).lower()
            
            # Ищем столбцы, содержащие слово "резюме"
            if 'резюме' in header_lower:
                resume_columns[i] = 'резюме'  # используем обобщенное название
        
        return resume_columns
