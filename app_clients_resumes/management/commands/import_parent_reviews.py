import os
import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.conf import settings
from app_clients_resumes.models import ParentReview


class Command(BaseCommand):
    help = 'Импорт отзывов родителей из Excel файлов'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            help='Путь к Excel файлу для импорта',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Показать что будет импортировано без сохранения в БД',
        )

    def handle(self, *args, **options):
        file_path = options.get('file')
        dry_run = options.get('dry_run', False)

        # Если файл не указан, ищем все Excel файлы в static/xlsx-files/
        if not file_path:
            xlsx_dir = os.path.join(settings.BASE_DIR, 'static', 'xlsx-files')
            if not os.path.exists(xlsx_dir):
                raise CommandError(f'Папка {xlsx_dir} не существует')
            
            excel_files = [f for f in os.listdir(xlsx_dir) if f.endswith('.xlsx')]
            if not excel_files:
                raise CommandError(f'В папке {xlsx_dir} не найдено Excel файлов')
            
            self.stdout.write(f'Найдено {len(excel_files)} Excel файлов для обработки')
            
            for excel_file in excel_files:
                file_full_path = os.path.join(xlsx_dir, excel_file)
                self.process_excel_file(file_full_path, dry_run)
        else:
            if not os.path.exists(file_path):
                raise CommandError(f'Файл {file_path} не существует')
            self.process_excel_file(file_path, dry_run)

        if dry_run:
            self.stdout.write(
                self.style.WARNING('Это был пробный запуск. Данные не были сохранены в БД.')
            )
        else:
            self.stdout.write(
                self.style.SUCCESS('Импорт отзывов родителей завершен успешно!')
            )

    def process_excel_file(self, file_path, dry_run):
        """Обрабатывает один Excel файл"""
        self.stdout.write(f'\nОбработка файла: {os.path.basename(file_path)}')
        
        try:
            workbook = openpyxl.load_workbook(file_path)
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f'Ошибка при открытии файла {file_path}: {e}')
            )
            return

        # Обрабатываем все листы
        for sheet_name in workbook.sheetnames:
            self.stdout.write(f'  Обработка листа: {sheet_name}')
            worksheet = workbook[sheet_name]
            self.process_worksheet(worksheet, dry_run, sheet_name)

    def process_worksheet(self, worksheet, dry_run, sheet_name):
        """Обрабатывает один лист Excel файла"""
        if worksheet.max_row < 2:
            self.stdout.write(f'    Лист {sheet_name} пуст или содержит только заголовки')
            return

        # Получаем заголовки из первой строки
        headers = []
        for col in range(1, worksheet.max_column + 1):
            header = worksheet.cell(row=1, column=col).value
            headers.append(header)

        self.stdout.write(f'    Найдено столбцов: {len(headers)}')
        self.stdout.write(f'    Заголовки: {headers[:8]}...')  # Показываем первые 8

        # Определяем индексы столбцов с отзывами родителей
        review_columns = self.identify_review_columns(headers)
        if not review_columns:
            self.stdout.write(f'    В листе {sheet_name} не найдено столбцов с отзывами родителей')
            return

        self.stdout.write(f'    Найдено столбцов с отзывами: {len(review_columns)}')

        # Обрабатываем каждую строку данных
        processed_count = 0
        saved_count = 0
        
        for row_num in range(2, worksheet.max_row + 1):
            # Получаем ID ребенка (первый столбец)
            student_id_cell = worksheet.cell(row=row_num, column=1)
            if student_id_cell.value is None:
                continue
            
            # Правильно обрабатываем числовые ID (убираем .0 если это целое число)
            if isinstance(student_id_cell.value, (int, float)):
                if float(student_id_cell.value).is_integer():
                    student_crm_id = str(int(student_id_cell.value))
                else:
                    student_crm_id = str(student_id_cell.value)
            else:
                student_crm_id = str(student_id_cell.value).strip()
            
            if not student_crm_id or student_crm_id == 'None':
                continue

            # Получаем ФИО ребенка (второй столбец) для логирования
            student_name_cell = worksheet.cell(row=row_num, column=2)
            student_name = str(student_name_cell.value).strip() if student_name_cell.value else 'Неизвестно'

            processed_count += 1

            # Обрабатываем каждый столбец с отзывом
            for col_index, review_type in review_columns.items():
                review_cell = worksheet.cell(row=row_num, column=col_index)
                review_content = review_cell.value
                
                # Пропускаем пустые отзывы
                if not review_content or str(review_content).strip() in ['', 'None']:
                    continue

                review_content = str(review_content).strip()
                
                if dry_run:
                    self.stdout.write(
                        f'    [DRY RUN] Сохранил бы отзыв: ID={student_crm_id}, '
                        f'Имя={student_name}, Тип={review_type}, '
                        f'Длина={len(review_content)} символов'
                    )
                else:
                    # Сохраняем отзыв в БД
                    review, created = ParentReview.objects.get_or_create(
                        student_crm_id=student_crm_id,
                        defaults={
                            'content': review_content
                        }
                    )
                    
                    if not created:
                        # Обновляем существующий отзыв
                        review.content = review_content
                        review.save()
                        action = 'обновлен'
                    else:
                        action = 'создан'
                    
                    saved_count += 1
                    self.stdout.write(
                        f'    Отзыв {action}: ID={student_crm_id}, '
                        f'Имя={student_name}'
                    )

        self.stdout.write(
            f'    Обработано строк: {processed_count}, '
            f'Сохранено отзывов: {saved_count}'
        )

    def identify_review_columns(self, headers):
        """Определяет какие столбцы содержат отзывы родителей"""
        review_columns = {}
        
        for i, header in enumerate(headers, 1):
            if not header:
                continue
                
            header_lower = str(header).lower()
            
            # Ищем столбцы, содержащие слова, связанные с отзывами родителей
            if any(keyword in header_lower for keyword in ['отзыв', 'родитель', 'мама', 'папа', 'parent', 'review']):
                review_columns[i] = str(header).strip()
        
        return review_columns
